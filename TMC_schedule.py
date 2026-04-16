import requests
import pandas as pd
import time
import re
import os
import getpass
from collections import Counter

# =============================
# ⚙️ CONFIGURATION
# =============================
BASE = "https://api.eu.cloud.talend.com"

def _load_token():
    """Charge le token depuis la variable d'env TMC_TOKEN, un fichier .env, ou la saisie interactive."""
    # 1. Variable d'environnement
    token = os.environ.get("TMC_TOKEN", "").strip()
    if token and not token.startswith("#"):
        return token

    # 2. Fichier .env dans le répertoire courant
    env_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.isfile(env_file):
        with open(env_file) as f:
            for line in f:
                line = line.strip()
                if line.startswith("TMC_TOKEN="):
                    token = line.split("=", 1)[1].strip().strip('"').strip("'")
                    if token:
                        return token

    # 3. Saisie interactive (masquée)
    print("⚠️  Token TMC non trouvé dans TMC_TOKEN ou .env")
    token = getpass.getpass("🔑 Entre ton token Talend Cloud (Bearer) : ").strip()
    if not token:
        raise SystemExit("❌ Token vide — arrêt du script.")
    return token

TOKEN = _load_token()
HEADERS = {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}


def _validate_token():
    """Vérifie que le token est valide en faisant un appel léger à l'API."""
    print("🔐 Vérification du token...")
    resp = requests.get(f"{BASE}/orchestration/projects?limit=1", headers=HEADERS)
    if resp.status_code == 401:
        raise SystemExit("❌ Token invalide (401 Unauthorized). Vérifie ton token et relance le script.")
    if not resp.ok:
        raise SystemExit(f"❌ Erreur API ({resp.status_code}): {resp.text}")
    print("✅ Token valide.\n")

_validate_token()


# =============================
# 🧰 HTTP UTIL
# =============================
def http_get(url):
    print(f"[GET] {url}")
    resp = requests.get(url, headers=HEADERS)
    if not resp.ok:
        print(f"⚠️ {resp.status_code}: {resp.text}")
        return None
    return resp.json()


# =============================
# 📁 PROJETS
# =============================
def fetch_projects():
    print("\n🔍 Récupération des projets Talend Studio...")
    data = http_get(f"{BASE}/orchestration/projects?limit=100")
    if not data:
        return {}
    projects = data.get("items", []) if isinstance(data, dict) else data
    project_map = {p["id"]: p.get("name") or p.get("technicalLabel") for p in projects if p.get("id")}
    print(f"✅ {len(projects)} projets détectés.")
    return project_map


# =============================
# 🌍 WORKSPACES
# =============================
def fetch_workspaces():
    print("\n🔍 Récupération des workspaces (v2.6 endpoint global)...")
    ws_data = http_get(f"{BASE}/orchestration/workspaces?limit=200")
    if not ws_data:
        return []
    ws_items = ws_data.get("items", []) if isinstance(ws_data, dict) else (ws_data or [])
    workspaces = []
    for w in ws_items:
        env_info = w.get("environment", {})
        workspaces.append({
            "workspace_id": w.get("id"),
            "workspace_name": w.get("name"),
            "environment_id": env_info.get("id"),
            "environment_name": env_info.get("name")
        })
    print(f"✅ {len(workspaces)} workspaces trouvés au total.")
    return workspaces


# =============================
# ⏰ SCHEDULES
# =============================
def fetch_schedules(environment_id):
    all_data, offset = [], 0
    print(f"\n🔄 Récupération des schedules ({environment_id})...")
    while True:
        url = f"{BASE}/orchestration/schedules?limit=100&offset={offset}&environmentId={environment_id}"
        data = http_get(url)
        items = (data or {}).get("items", []) if isinstance(data, dict) else (data or [])
        if not items:
            break
        all_data.extend(items)
        if len(items) < 100:
            break
        offset += 100
        time.sleep(0.2)
    print(f"✅ {len(all_data)} schedules récupérées.")
    return all_data


# =============================
# ⚙️ TASKS
# =============================
def fetch_tasks(environment_id):
    all_data, offset = [], 0
    print(f"\n🔄 Récupération des tasks ({environment_id})...")
    while True:
        url = f"{BASE}/orchestration/executables/tasks?limit=100&offset={offset}&environmentId={environment_id}"
        data = http_get(url)
        items = (data or {}).get("items", []) if isinstance(data, dict) else (data or [])
        if not items:
            break
        all_data.extend(items)
        if len(items) < 100:
            break
        offset += 100
        time.sleep(0.2)
    print(f"✅ {len(all_data)} tasks récupérées.")
    return all_data


# =============================
# 📦 PLANS
# =============================
def fetch_plans(environment_id):
    all_data, offset = [], 0
    print(f"\n🔄 Récupération des plans ({environment_id})...")
    while True:
        url = f"{BASE}/orchestration/executables/plans?limit=100&offset={offset}&environmentId={environment_id}"
        data = http_get(url)
        items = (data or {}).get("items", []) if isinstance(data, dict) else (data or [])
        if not items:
            break
        all_data.extend(items)
        if len(items) < 100:
            break
        offset += 100
        time.sleep(0.2)
    print(f"✅ {len(all_data)} plans récupérés.")
    return all_data


def get_plan_task_ids(plan):
    """Extrait tous les IDs de tâches contenues dans un plan (parcours de la liste chaînée chart)."""
    task_ids = set()
    step = plan.get("chart", {})
    while step and isinstance(step, dict):
        for flow in step.get("flows", []):
            fid = flow.get("id")
            if fid:
                task_ids.add(fid)
        step = step.get("nextStep")
    return task_ids


# =============================
# 📦 ARTEFACTS
# =============================

# =============================
# 📦 ARTEFACTS
# =============================
def fetch_artifacts(workspace_id):
    all_data, offset = [], 0
    print(f"\n🔄 Artefacts du workspace {workspace_id}...")
    while True:
        url = f"{BASE}/orchestration/artifacts?limit=100&offset={offset}&workspaceId={workspace_id}"
        data = http_get(url)
        items = (data or {}).get("items", []) if isinstance(data, dict) else (data or [])
        if not items:
            break
        all_data.extend(items)
        if len(items) < 100:
            break
        offset += 100
        time.sleep(0.2)
    print(f"✅ {len(all_data)} artefacts récupérés.")
    return all_data


# =============================
# 🧩 MAPPING WORKSPACE → PROJET
# =============================
def map_workspaces_to_projects(workspaces, project_map):
    workspace_project_map = {}
    print("\n🧩 Correspondance Workspaces ↔ Projets Talend")
    print("--------------------------------------------------")
    for w in workspaces:
        ws_name = (w["workspace_name"] or "").upper()
        env_name = (w["environment_name"] or "")
        matched_project = None
        for pname in project_map.values():
            if pname and ws_name in pname.upper():
                matched_project = pname
                break
        workspace_project_map[w["workspace_id"]] = matched_project or "Inconnu"
        print(f"Workspace: {w['workspace_name']:<12} | Env: {env_name:<6} | Projet: {matched_project or '❌ Aucun'}")
    print("--------------------------------------------------\n")
    return workspace_project_map


# =============================
# 🕒 LECTURE CRON (v3.5 améliorée)
# =============================

def _uniform_step(int_list):
    """Retourne le pas si la liste a un écart constant, sinon None."""
    if len(int_list) < 2:
        return None
    diffs = [b - a for a, b in zip(int_list, int_list[1:])]
    return diffs[0] if all(d == diffs[0] for d in diffs) else None

def _dow_label(dow_raw: str) -> str:
    d = dow_raw.upper().replace(",", ", ")
    if d in ("1-5", "2-6", "MON-FRI"):
        return "jours de semaine"
    if d in ("*", "1-7", "0-7", "0,1,2,3,4,5,6,7", "0,1,2,3,4,5,6", "1,2,3,4,5,6,7"):
        return "tous les jours"
    return d

def readable_cron(expr: str) -> str:
    if not expr:
        return ""
    expr = expr.strip()

    # ----- Normalisation : champ année wildcard (6 champs → 5 champs) -----
    # ex: "30 4 ? * 2-6 *" → "30 4 ? * 2-6"
    parts = expr.split()
    if len(parts) == 6 and parts[-1] in ('*', '?'):
        expr = ' '.join(parts[:5])

    # ----- Normalisation : mois 1-12 ou 1,2,...,12 → * -----
    parts = expr.split()
    if len(parts) == 5 and parts[3] in ('1-12', '1,2,3,4,5,6,7,8,9,10,11,12'):
        parts[3] = '*'
        expr = ' '.join(parts)

    # ----- Normalisation : dom=* mois=* dow=? → dom=? mois=* dow=* -----
    # ex: "00 20 * * ?"  /  "3,8,13 * * * ?" → formes gérées par les regex suivantes
    parts = expr.split()
    if len(parts) == 5 and parts[2] == '*' and parts[3] == '*' and parts[4] == '?':
        parts[2] = '?'
        parts[4] = '*'
        expr = ' '.join(parts)

    # ----- Cas spéciaux Talend : fin / jour ouvré de mois -----
    if re.search(r"\bL\b", expr) and not re.search(r"\bL[-W]", expr):
        # ex. "5 0 L * ?" → le dernier jour du mois à 00:05
        m = re.match(r"^(\d{1,2})\s+(\d{1,2})\s+L\s+\*\s+\?", expr)
        if m:
            minute, hour = map(int, m.groups())
            return f"Le dernier jour du mois à {hour:02d}h{minute:02d}"
        return "Le dernier jour du mois"
    if re.search(r"\bL-1\b", expr):
        # ex. "5 0 L-1 * ?" → le dernier jour du mois à 00:05
        m = re.match(r"^(\d{1,2})\s+(\d{1,2})\s+L-1\s+\*\s+\?", expr)
        if m:
            minute, hour = map(int, m.groups())
            return f"Le dernier jour du mois à {hour:02d}h{minute:02d}"
        return "Le dernier jour du mois"
    if re.search(r"\bLW\b", expr):
        # ex. "5 0 LW * ?" → dernier jour ouvré du mois à 00:05
        m = re.match(r"^(\d{1,2})\s+(\d{1,2})\s+LW\s+\*\s+\?", expr)
        if m:
            minute, hour = map(int, m.groups())
            return f"Le dernier jour ouvré du mois à {hour:02d}h{minute:02d}"
        return "Le dernier jour ouvré du mois"
    if re.search(r"\b1W\b", expr):
        # ex. "5 0 1W * ?" → 1er jour ouvré du mois à 00:05
        m = re.match(r"^(\d{1,2})\s+(\d{1,2})\s+1W\s+\*\s+\?", expr)
        if m:
            minute, hour = map(int, m.groups())
            return f"Le 1er jour ouvré du mois à {hour:02d}h{minute:02d}"
        return "Le 1er jour ouvré du mois"

    # ----- Toutes les X minutes sur une plage horaire (dow optionnel) -----
    # ex: "*/15 7-18 ? * 2,3,4,5,6"  /  "*/17 8-19 ? * 2-6"
    m = re.match(r"^\*/(\d+)\s+(\d{1,2})(?:-(\d{1,2}))?\s+\?\s+\*\s+([0-9A-Z,\-*]+)$", expr)
    if m:
        freq, h1, h2, dow = m.groups()
        h1 = int(h1); h2 = int(h2) if h2 else h1
        lbl = _dow_label(dow)
        if lbl == "jours de semaine":
            return f"Toutes les {int(freq)} min de {h1:02d}h à {h2:02d}h (jours de semaine)"
        if lbl == "tous les jours":
            return f"Toutes les {int(freq)} min de {h1:02d}h à {h2:02d}h"
        return f"Toutes les {int(freq)} min de {h1:02d}h à {h2:02d}h ({lbl})"

    # ----- Liste de minutes + plage d'heures -----
    # ex: "0,30 8-17 ? * 2,3,4,5,6"  → toutes les 30 min 08–17h (jours de semaine)
    m = re.match(r"^([\d,]+)\s+(\d{1,2})-(\d{1,2})\s+\?\s+\*\s+([0-9A-Z,\-*]+)$", expr)
    if m:
        minutes_csv, h1, h2, dow = m.groups()
        mins = [int(x) for x in minutes_csv.split(",") if x.isdigit()]
        step = _uniform_step(sorted(mins))
        lbl = _dow_label(dow)
        if step and set(mins) == set(range(0, 60, step)):
            base = f"Toutes les {step} min"
        elif step:
            base = f"Chaque {step} min (aux minutes {minutes_csv})"
        else:
            base = f"Aux minutes {minutes_csv}"
        suffix = "" if lbl in ("tous les jours",) else f" ({lbl})"
        return f"{base} de {int(h1):02d}h à {int(h2):02d}h{suffix}"

    # ----- Liste de minutes + liste/plage d'heures (générique) -----
    # ex: "0,10,20,30,40,50 5,6,7,8,9,10,... ? * *"
    m = re.match(r"^([\d,]+)\s+([\d,\-]+)\s+\?\s+\*\s+([0-9A-Z,\-*]+)$", expr)
    if m:
        minutes_csv, hours_field, dow = m.groups()
        mins = [int(x) for x in minutes_csv.split(",") if x.isdigit()]
        step = _uniform_step(sorted(mins))
        if "-" in hours_field:
            h1, h2 = [int(x) for x in hours_field.split("-")]
            hours_lbl = f"de {h1:02d}h à {h2:02d}h"
        else:
            hours = [int(h) for h in hours_field.split(",") if h.isdigit()]
            hours_lbl = "à " + ", ".join(f"{h:02d}h" for h in hours)
        lbl = _dow_label(dow)
        if step and set(mins) == set(range(0, 60, step)):
            base = f"Toutes les {step} min"
        elif step:
            base = f"Chaque {step} min (aux minutes {minutes_csv})"
        else:
            base = f"Aux minutes {minutes_csv}"
        suffix = "" if lbl in ("tous les jours",) else f" ({lbl})"
        return f"{base} {hours_lbl}{suffix}"

    # ----- Multi-heures (minute fixe) -----
    # ex: "0 9,14,21 ? * *"  /  "15 8,12,16 ? * 1-5"
    m = re.match(r"^(\d{1,2})\s+([\d,]+)\s+\?\s+\*\s+([0-9A-Z,\-*]+)$", expr)
    if m:
        minute, hours_csv, dow = m.groups()
        hours = [int(h) for h in hours_csv.split(",") if h.isdigit()]
        hour_str = ", ".join(f"{h:02d}h" for h in hours)
        lbl = _dow_label(dow)
        prefix = "Tous les jours" if lbl == "tous les jours" else ("Tous les jours de semaine" if lbl == "jours de semaine" else f"Chaque semaine ({lbl})")
        return f"{prefix} à {hour_str}{'' if int(minute)==0 else f' ({int(minute)}m)'}"

    # ----- Horaire unique (dow/hebdo) -----
    # ex: "20 5 ? * 2-6" / "0 2 ? * *"
    m = re.match(r"^(\d{1,2})\s+(\d{1,2})\s+\?\s+\*\s+([0-9A-Z,\-*]+)$", expr)
    if m:
        minute, hour, dow = m.groups()
        lbl = _dow_label(dow)
        if lbl == "jours de semaine":
            return f"Tous les jours de semaine à {int(hour):02d}h{int(minute):02d}"
        if lbl == "tous les jours":
            return f"Tous les jours à {int(hour):02d}h{int(minute):02d}"
        return f"Chaque semaine ({lbl}) à {int(hour):02d}h{int(minute):02d}"

    # ----- Jours du mois + mois (exécutions mensuelles/annuelles) -----
    # ex: "00 8 1-31 1-12 ?" / "10 8 1-31 1-12 ?" / "00 10 29 3,6,9,12 ? *"
    m = re.match(r"^(\d{1,2})\s+(\d{1,2})\s+([\d,\-]+)\s+([\d,\-*]+)\s+\?\s*\*?$", expr)
    if m:
        minute, hour, dom, months = m.groups()
        if months in ("*", "1-12"):
            return f"Chaque mois, jours {dom}, à {int(hour):02d}h{int(minute):02d}"
        return f"Jours {dom} des mois {months} à {int(hour):02d}h{int(minute):02d}"

    # ----- Récurrences simples sans plage -----
    # ex: "*/5 * * * * ?" / "*/10 * * * * ?"
    if re.search(r"\*/(\d+)", expr) and "?" in expr:
        freq = int(re.search(r"\*/(\d+)", expr).group(1))
        return f"Toutes les {freq} minutes"

    # ----- Minute fixe, toutes les heures (hour=*) -----
    # ex: "0 * ? * *" / "15 * ? * *"
    m = re.match(r"^(\d{1,2})\s+\*\s+\?\s+\*\s+([0-9A-Z,\-*]+)$", expr)
    if m:
        minute, dow = m.groups()
        lbl = _dow_label(dow)
        suffix = "" if lbl == "tous les jours" else f" ({lbl})"
        if int(minute) == 0:
            return f"Toutes les heures{suffix}"
        return f"Toutes les heures (à :{int(minute):02d}){suffix}"

    # ----- Liste de minutes, toutes les heures (hour=*) -----
    # ex: "0,5,10,...,55 * ? * *" / "0,15,30,45 * ? * *"
    m = re.match(r"^([\d,]+)\s+\*\s+\?\s+\*\s+([0-9A-Z,\-*]+)$", expr)
    if m:
        minutes_csv, dow = m.groups()
        mins = sorted([int(x) for x in minutes_csv.split(",") if x.isdigit()])
        step = _uniform_step(mins)
        lbl = _dow_label(dow)
        if step and set(mins) == set(range(0, 60, step)):
            base = f"Toutes les {step} min"
        else:
            base = f"Aux minutes {minutes_csv}"
        suffix = "" if lbl == "tous les jours" else f" ({lbl})"
        return f"{base} (24h/24){suffix}"

    # ----- Jours du mois spécifiques + heures multiples -----
    # ex: "0 12,18 1,2 * ?"
    m = re.match(r"^(\d{1,2})\s+([\d,]+)\s+([\d,]+)\s+\*\s+\?$", expr)
    if m:
        minute, hours_csv, dom = m.groups()
        hours = [int(h) for h in hours_csv.split(",") if h.isdigit()]
        hour_str = ", ".join(f"{h:02d}h" for h in hours)
        min_suffix = "" if int(minute) == 0 else f" ({int(minute)}m)"
        return f"Les jours {dom} du mois à {hour_str}{min_suffix}"

    # ----- Fallback : on remonte l’expression pour debug -----
    return f"(CRON: {expr})"

# =============================
# 🔢 CLASSIFICATION
# =============================
def classify_schedule(desc):
    dl = desc.lower()
    if "5 min" in dl:
        return "Every 5min"
    if "10 min" in dl:
        return "Every 10min"
    if re.search(r"toutes les \d+ min", dl):
        return "Recurring"
    if "aux minutes" in dl or ("chaque" in dl and "min" in dl):
        return "Recurring"
    if "mois" in dl or "mensuel" in dl:
        return "Monthly"
    if "tous les jours" in dl or "quotidien" in dl:
        return "Daily"
    if "semaine" in dl:
        return "Weekly"
    return "Others"


def hour_from_cron(expr):
    try:
        parts = expr.strip().split()
        if len(parts) >= 2:
            h = parts[1]
            m = re.match(r"(\d+)", h)
            if m:
                return int(m.group(1))
    except Exception:
        pass
    return None


# =============================
# 📊 BUILD DATAFRAME
# =============================
def _resolve_plan_project(plan, workspace_project_map):
    """Résout le projet d'un plan via son workspace."""
    plan_ws = plan.get("workspace") or {}
    ws_id = plan_ws.get("id", "")
    project = workspace_project_map.get(ws_id)
    if project and project != "Inconnu":
        return project
    # Fallback : utiliser le nom du workspace directement
    return plan_ws.get("name", "Inconnu")


def schedule_map_for(schedules, exec_id):
    """Retourne les infos de schedule pour un exécutable donné, ou None."""
    for s in schedules:
        if s.get("executableId") == exec_id:
            triggers = s.get("triggers", [])
            cron_exprs = [trig.get("cronExpression", "") for trig in triggers if trig.get("cronExpression")]
            descs = [readable_cron(c) for c in cron_exprs]
            hours = [hour_from_cron(c) for c in cron_exprs]

            pause_info = s.get("planPauseDetails") or {}
            is_paused = pause_info.get("pause", False)
            if is_paused:
                status = "⏸️ En pause"
            elif cron_exprs:
                status = "✅ Planifié (via Plan)"
            else:
                status = "❌ Manuel / Inactif"

            return {"cron_exprs": cron_exprs, "descs": descs, "hours": hours, "status": status}
    return None


def build_dataframe(artifacts, schedules, tasks, plans, workspace_id, workspace_project_map):
    task_rows = []
    plan_rows = []
    artifact_map = {a["id"]: a for a in artifacts}
    task_map = {t.get("id") or t.get("executable"): t for t in tasks}
    plan_map = {p.get("id") or p.get("executable"): p for p in plans}

    valid_artifact_ids = {a["id"] for a in artifacts}

    # Construire la correspondance tâche → plan (nom + schedule), uniquement pour les plans du workspace courant
    task_to_plan = {}  # task_id → {"plan_name", "schedule"}
    for plan in plans:
        plan_ws_id = (plan.get("workspace") or {}).get("id", "")
        if plan_ws_id != workspace_id:
            continue
        plan_id = plan.get("id") or plan.get("executable")
        plan_schedule = schedule_map_for(schedules, plan_id)
        plan_task_ids = get_plan_task_ids(plan)
        for tid in plan_task_ids:
            task_to_plan[tid] = {
                "plan_name": plan.get("name", ""),
                "schedule": plan_schedule,
            }

    # Associer chaque schedule à son exécutable
    schedule_map = {}
    for s in schedules:
        exec_id = s.get("executableId")
        if exec_id:
            schedule_map[exec_id] = s

    seen_task_ids = set()
    seen_plan_ids = set()

    # ── TASKS avec leur propre schedule ──
    for s in schedules:
        exec_id = s.get("executableId")
        if not exec_id or exec_id in seen_task_ids:
            continue
        if exec_id in plan_map:
            continue

        task = task_map.get(exec_id, {})
        if not task:
            continue
        if task.get("artifactId") not in valid_artifact_ids:
            continue
        seen_task_ids.add(exec_id)

        artifact = artifact_map.get(task.get("artifactId") or "", {})
        ws = (artifact.get("workspace") or {}).get("id", workspace_id)
        project_name = workspace_project_map.get(ws, "Inconnu")
        environment = ((artifact.get("workspace") or {}).get("environment") or {}).get("name", "")

        triggers = s.get("triggers", [])
        cron_exprs = [trig.get("cronExpression", "") for trig in triggers if trig.get("cronExpression")]
        descs = [readable_cron(c) for c in cron_exprs]
        categories = [classify_schedule(d) for d in descs]
        hours = [hour_from_cron(c) for c in cron_exprs]

        pause_info = task.get("taskPauseDetails") or {}
        is_paused = pause_info.get("pause", False)
        if is_paused:
            status = "⏸️ En pause"
        elif cron_exprs:
            status = "✅ Planifié"
        else:
            status = "❌ Manuel / Inactif"

        category = max(set(categories), key=categories.count) if categories else "Others"

        # Vérifier si la tâche fait aussi partie d'un plan
        plan_info = task_to_plan.get(exec_id)

        task_rows.append({
            "Projet": project_name,
            "Nom": task.get("name", ""),
            "Job / Artefact": artifact.get("name", ""),
            "Plan": plan_info["plan_name"] if plan_info else "",
            "Expression CRON": " | ".join(cron_exprs),
            "Description": " | ".join(descs),
            "Catégorie": category,
            "Heure": hours[0] if hours else None,
            "Statut": status,
            "Environnement": environment,
        })

    # ── TASKS sans leur propre schedule ──
    for task_id, task in task_map.items():
        if task_id in seen_task_ids:
            continue
        if task.get("artifactId") not in valid_artifact_ids:
            continue
        seen_task_ids.add(task_id)

        artifact = artifact_map.get(task.get("artifactId") or "", {})
        ws = (artifact.get("workspace") or {}).get("id", workspace_id)
        project_name = workspace_project_map.get(ws, "Inconnu")
        environment = ((artifact.get("workspace") or {}).get("environment") or {}).get("name", "")

        # Si la tâche fait partie d'un plan, hériter du schedule du plan
        plan_info = task_to_plan.get(task_id)
        if plan_info and plan_info["schedule"]:
            ps = plan_info["schedule"]
            cron_exprs = ps["cron_exprs"]
            descs = ps["descs"]
            categories = [classify_schedule(d) for d in descs]
            hours = [hour_from_cron(c) for c in cron_exprs]
            category = max(set(categories), key=categories.count) if categories else "Others"
            status = ps["status"]
        else:
            cron_exprs, descs, hours = [], [], []
            category = "Others"
            status = "❌ Manuel / Inactif"

        task_rows.append({
            "Projet": project_name,
            "Nom": task.get("name", ""),
            "Job / Artefact": artifact.get("name", ""),
            "Plan": plan_info["plan_name"] if plan_info else "",
            "Expression CRON": " | ".join(cron_exprs),
            "Description": " | ".join(descs),
            "Catégorie": category,
            "Heure": hours[0] if hours else None,
            "Statut": status,
            "Environnement": environment,
        })

    # ── PLANS avec schedule ──
    for s in schedules:
        exec_id = s.get("executableId")
        if not exec_id or exec_id in seen_plan_ids:
            continue
        plan = plan_map.get(exec_id)
        if not plan:
            continue
        # Filtrer : ne garder que les plans du workspace courant
        plan_ws_id = (plan.get("workspace") or {}).get("id", "")
        if plan_ws_id != workspace_id:
            continue
        seen_plan_ids.add(exec_id)

        project_name = _resolve_plan_project(plan, workspace_project_map)
        plan_env = (plan.get("workspace") or {}).get("environment", {}).get("name", "")

        plan_tasks = get_plan_task_ids(plan)
        task_names = [task_map.get(tid, {}).get("name", tid) for tid in plan_tasks]

        triggers = s.get("triggers", [])
        cron_exprs = [trig.get("cronExpression", "") for trig in triggers if trig.get("cronExpression")]
        descs = [readable_cron(c) for c in cron_exprs]
        categories = [classify_schedule(d) for d in descs]
        hours = [hour_from_cron(c) for c in cron_exprs]

        pause_info = plan.get("planPauseDetails") or {}
        is_paused = pause_info.get("pause", False)
        if is_paused:
            status = "⏸️ En pause"
        elif cron_exprs:
            status = "✅ Planifié"
        else:
            status = "❌ Manuel / Inactif"

        category = max(set(categories), key=categories.count) if categories else "Others"

        plan_rows.append({
            "Projet": project_name,
            "Nom": plan.get("name", ""),
            "Tâches incluses": ", ".join(task_names) if task_names else "",
            "Expression CRON": " | ".join(cron_exprs),
            "Description": " | ".join(descs),
            "Catégorie": category,
            "Heure": hours[0] if hours else None,
            "Statut": status,
            "Environnement": plan_env,
        })

    # ── PLANS sans schedule ──
    for plan_id, plan in plan_map.items():
        if plan_id in seen_plan_ids:
            continue
        # Filtrer : ne garder que les plans du workspace courant
        plan_ws_id = (plan.get("workspace") or {}).get("id", "")
        if plan_ws_id != workspace_id:
            continue
        seen_plan_ids.add(plan_id)

        project_name = _resolve_plan_project(plan, workspace_project_map)
        plan_env = (plan.get("workspace") or {}).get("environment", {}).get("name", "")

        plan_tasks = get_plan_task_ids(plan)
        task_names = [task_map.get(tid, {}).get("name", tid) for tid in plan_tasks]

        plan_rows.append({
            "Projet": project_name,
            "Nom": plan.get("name", ""),
            "Tâches incluses": ", ".join(task_names) if task_names else "",
            "Expression CRON": "",
            "Description": "",
            "Catégorie": "Others",
            "Heure": None,
            "Statut": "❌ Manuel / Inactif",
            "Environnement": plan_env,
        })

    df_tasks = pd.DataFrame(task_rows)
    df_plans = pd.DataFrame(plan_rows)
    return df_tasks, df_plans


# =============================
# 📈 EXPORT
# =============================
def _style_table(ws, df_len, df_cols, table_name, start_row=1):
    """Applique un style tableau Excel + auto-largeur des colonnes."""
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    end_col = get_column_letter(len(df_cols))
    end_row = start_row + df_len
    ref = f"A{start_row}:{end_col}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    # Auto-largeur
    for i, col in enumerate(df_cols, 1):
        max_len = len(str(col))
        for row in range(start_row + 1, end_row + 1):
            val = ws.cell(row=row, column=i).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 60)


def export_excel(df_tasks, df_plans, project, env):
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint

    safe_project = re.sub(r'[^\w\-]', '_', project)
    safe_env = re.sub(r'[^\w\-]', '_', env)
    output = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"tmc_schedules_{safe_project}_{safe_env}.xlsx")
    print(f"\n💾 Génération du fichier Excel : {output}")

    # Combiner tasks + plans pour les stats globales
    df_all = pd.concat([df_tasks, df_plans], ignore_index=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ── Onglet Tasks ──
        if not df_tasks.empty:
            tasks_sorted = df_tasks.sort_values(by=["Statut", "Nom"])
            tasks_sorted.to_excel(writer, sheet_name="Tasks", index=False)
            _style_table(writer.sheets["Tasks"], len(tasks_sorted), tasks_sorted.columns, "TblTasks")

        # ── Onglet Plans ──
        if not df_plans.empty:
            plans_sorted = df_plans.sort_values(by=["Statut", "Nom"])
            plans_sorted.to_excel(writer, sheet_name="Plans", index=False)
            _style_table(writer.sheets["Plans"], len(plans_sorted), plans_sorted.columns, "TblPlans")

        # ── Onglet Recap (catégories) ──
        recap = df_all["Catégorie"].value_counts().reset_index()
        recap.columns = ["Catégorie", "Nombre"]
        recap.loc[len(recap)] = ["Total", recap["Nombre"].sum()]
        recap.to_excel(writer, sheet_name="Recap", index=False)
        ws_recap = writer.sheets["Recap"]
        _style_table(ws_recap, len(recap), recap.columns, "TblRecap")

        pie = PieChart()
        pie.title = "Répartition par catégorie"
        pie.style = 10
        pie.width = 16
        pie.height = 10
        cats = Reference(ws_recap, min_col=1, min_row=2, max_row=1 + len(recap) - 1)
        vals = Reference(ws_recap, min_col=2, min_row=1, max_row=1 + len(recap) - 1)
        pie.add_data(vals, titles_from_data=True)
        pie.set_categories(cats)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True
        ws_recap.add_chart(pie, "D2")

        # ── Onglet Statuts ──
        stat = df_all["Statut"].value_counts().reset_index()
        stat.columns = ["Statut", "Nombre"]
        stat.loc[len(stat)] = ["Total", stat["Nombre"].sum()]
        stat.to_excel(writer, sheet_name="Statuts", index=False)
        ws_stat = writer.sheets["Statuts"]
        _style_table(ws_stat, len(stat), stat.columns, "TblStatuts")

        pie2 = PieChart()
        pie2.title = "Répartition par statut"
        pie2.style = 10
        pie2.width = 16
        pie2.height = 10
        cats2 = Reference(ws_stat, min_col=1, min_row=2, max_row=1 + len(stat) - 1)
        vals2 = Reference(ws_stat, min_col=2, min_row=1, max_row=1 + len(stat) - 1)
        pie2.add_data(vals2, titles_from_data=True)
        pie2.set_categories(cats2)
        pie2.dataLabels = DataLabelList()
        pie2.dataLabels.showPercent = True
        pie2.dataLabels.showVal = True

        color_map = {
            "✅ Planifié": "2ECC71",
            "❌ Manuel / Inactif": "E74C3C",
            "⏸️ En pause": "F39C12",
        }
        stat_labels = stat["Statut"].tolist()[:-1]
        for i, label in enumerate(stat_labels):
            color = color_map.get(label)
            if color:
                pt = DataPoint(idx=i)
                pt.graphicalProperties.solidFill = color
                pie2.series[0].data_points.append(pt)
        ws_stat.add_chart(pie2, "D2")

        # ── Onglet Affluence ──
        all_hours = []
        for cron_field in df_all["Expression CRON"].dropna():
            for cron_expr in str(cron_field).split(" | "):
                h = hour_from_cron(cron_expr)
                if h is not None:
                    all_hours.append(h)
        afflu = Counter(all_hours)
        afflu_df = pd.DataFrame({"Heure": range(24), "Nombre de déclenchements": [afflu.get(h, 0) for h in range(24)]})
        afflu_df.to_excel(writer, sheet_name="Affluence", index=False)
        ws_afflu = writer.sheets["Affluence"]
        _style_table(ws_afflu, len(afflu_df), afflu_df.columns, "TblAffluence")

        bar = BarChart()
        bar.type = "col"
        bar.title = "Affluence des déclenchements par heure"
        bar.y_axis.title = "Nombre de déclenchements"
        bar.x_axis.title = "Heure"
        bar.style = 10
        bar.width = 22
        bar.height = 12
        bar_cats = Reference(ws_afflu, min_col=1, min_row=2, max_row=25)
        bar_vals = Reference(ws_afflu, min_col=2, min_row=1, max_row=25)
        bar.add_data(bar_vals, titles_from_data=True)
        bar.set_categories(bar_cats)
        bar.legend = None
        bar.x_axis.tickLblPos = "low"
        bar.x_axis.delete = False
        bar.y_axis.tickLblPos = "low"
        bar.y_axis.delete = False
        bar.x_axis.numFmt = '0"h"'
        bar.y_axis.numFmt = '0'
        ws_afflu.add_chart(bar, "D2")

    print(f"✅ Export terminé : {output}\n")


# =============================
# 🚀 MAIN
# =============================
if __name__ == "__main__":
    print("🚀 Export Talend Cloud Schedules – NextDecision Edition v3.5\n")

    projects = fetch_projects()
    workspaces = fetch_workspaces()
    workspace_project_map = map_workspaces_to_projects(workspaces, projects)

    print("Sélectionne les workspaces à exporter :")
    for i, w in enumerate(workspaces, start=1):
        print(f"[{i}] {w['workspace_name']} – {w['environment_name']}")
    print("[*] Tous les workspaces")

    choice = input("\n👉 Entre les numéros séparés par des virgules (ex: 1,3) ou * pour tout : ").strip()
    if choice == "*":
        selected = workspaces
    else:
        try:
            idx = [int(x.strip()) for x in choice.split(",") if x.strip().isdigit()]
            selected = [workspaces[i - 1] for i in idx if 0 < i <= len(workspaces)]
        except Exception:
            selected = []

    print(f"\n🧭 {len(selected)} workspace(s) sélectionné(s).\n")

    for w in selected:
        project = w["workspace_name"]
        env = w["environment_name"]
        ws_id = w["workspace_id"]
        print(f"🏗️ Traitement du projet {project} – {env}...")

        artifacts = fetch_artifacts(ws_id)
        schedules = fetch_schedules(w["environment_id"])
        tasks = fetch_tasks(w["environment_id"])
        plans = fetch_plans(w["environment_id"])

        df_tasks, df_plans = build_dataframe(artifacts, schedules, tasks, plans, ws_id, workspace_project_map)
        if not df_tasks.empty or not df_plans.empty:
            print(f"  → {len(df_tasks)} tâches, {len(df_plans)} plans")
            export_excel(df_tasks, df_plans, project, env)
        else:
            print(f"⚠️ Aucune donnée trouvée pour {project} – {env}")
